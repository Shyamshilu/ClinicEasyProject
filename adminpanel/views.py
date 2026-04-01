from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q, Count
from accounts.models import Patient
from appointments import appointments
from .forms import PatientForm,AppointmentForm
from accounts import models
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from appointments.models import Appointment,ContactMessage
from doctors.models import Doctor
from adminpanel.forms import DoctorForm
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@csrf_exempt
def admin_login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_superuser:
            login(request, user)
            return redirect('admin_dashboard')
        else:
            return render(request, 'adminpanel/login.html', {
                'error': 'Invalid admin credentials'
            })

    return render(request, 'adminpanel/login.html')


@login_required(login_url='admin_login')
def admin_logout(request):
    logout(request)
    return redirect('admin_login')


def admin_dashboard(request):
    total_users = User.objects.filter(is_staff=False).count()
    total_admins = User.objects.filter(is_staff=True).count()
    total_appointments = Appointment.objects.count()

    approved_count = Appointment.objects.filter(status='Approved').count()
    pending_count = Appointment.objects.filter(status='Pending').count()
    rejected_count = Appointment.objects.filter(status='Rejected').count()
    cancelled_count = Appointment.objects.filter(status='Cancelled').count()

    status_data = {
        'Approved': approved_count,
        'Pending': pending_count,
        'Rejected': rejected_count,
        'Cancelled': cancelled_count,
    }

    context = {
        'total_users': total_users,
        'total_admins': total_admins,
        'total_appointments': total_appointments,
        'status_data': status_data,
    }

    return render(request, 'adminpanel/dashboard.html', context)

@login_required(login_url='admin_login')
def admin_users(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    if request.method == "POST":
        action = request.POST.get('action')
        user_ids = request.POST.getlist('user_ids')

        if user_ids:
            if action == "delete":
                User.objects.filter(id__in=user_ids).delete()

            elif action == "make_admin":
                User.objects.filter(id__in=user_ids).update(
                    is_superuser=True,
                    is_staff=True
                )

        return redirect('admin_users')

    search = request.GET.get('search')
    if search in (None, '', 'None'):
        search = ''

    role = request.GET.get('role')

    users = User.objects.all().order_by('-id')

    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search)
        )

    if role == 'admin':
        users = users.filter(is_staff=True)
    elif role == 'user':
        users = users.filter(is_staff=False)

    paginator = Paginator(users, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'adminpanel/users.html', {
        'page_obj': page_obj,
        'search': search
    })


@login_required(login_url='admin_login')
def admin_add_user(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    if request.method == "POST":
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        is_superuser = request.POST.get('is_superuser') == 'on'

        User.objects.create_user(
            username=username,
            email=email,
            password=password,
            is_superuser=is_superuser,
            is_staff=is_superuser
        )
        return redirect('admin_users')

    return render(request, 'adminpanel/add_user.html')


@login_required(login_url='admin_login')
def admin_edit_user(request, user_id):
    if not request.user.is_superuser:
        return redirect('admin_login')

    user = User.objects.get(id=user_id)

    if request.method == "POST":
        user.username = request.POST['username']
        user.email = request.POST['email']
        user.is_superuser = request.POST.get('is_superuser') == 'on'
        user.is_staff = user.is_superuser

        if request.POST.get('password'):
            user.set_password(request.POST['password'])

        user.save()
        return redirect('admin_users')

    return render(request, 'adminpanel/edit_user.html', {'user': user})


@login_required(login_url='admin_login')
def admin_delete_user(request, user_id):
    if not request.user.is_superuser:
        return redirect('admin_login')

    user = User.objects.get(id=user_id)

    if request.method == "POST":
        user.delete()
        return redirect('admin_users')

    return render(request, 'adminpanel/delete_user.html', {'user': user})



@login_required(login_url='admin_login')
def admin_patients(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    patients = Patient.objects.all().order_by('-id')

    return render(request, 'adminpanel/patients/list.html', {
        'patients': patients
    })


@login_required(login_url='admin_login')
def admin_add_patient(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    form = PatientForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('admin_patients')

    return render(request, 'adminpanel/patients/add.html', {
        'form': form
    })

@login_required(login_url='admin_login')
def admin_edit_patient(request, patient_id):
    if not request.user.is_superuser:
        return redirect('admin_login')

    patient = Patient.objects.get(id=patient_id)
    form = PatientForm(request.POST or None, instance=patient)

    if form.is_valid():
        form.save()
        return redirect('admin_patients')

    return render(request, 'adminpanel/patients/edit.html', {
        'form': form,
        'patient': patient
    })


@login_required(login_url='admin_login')
def admin_delete_patient(request, patient_id):
    if not request.user.is_superuser:
        return redirect('admin_login')

    patient = Patient.objects.get(id=patient_id)

    if request.method == "POST":
        patient.delete()
        return redirect('admin_patients')

    return render(request, 'adminpanel/patients/delete.html', {
        'patient': patient
    })

@login_required(login_url='admin_login')
def admin_appointments(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    appointments = Appointment.objects.select_related(
        'doctor', 'patient'
    ).order_by('-appointment_date', '-appointment_time')

    return render(request, 'adminpanel/appointments/list.html', {
        'appointments': appointments
    })


def admin_add_appointment(request):
    if request.method == 'POST':
        form = AppointmentForm(request.POST)

        if form.is_valid(): 
            form.save()
            return redirect('admin_appointments')
        else:
            print(form.errors)  
    else:
        form = AppointmentForm()

    return render(request, 'adminpanel/appointments/add.html', {
        'form': form
    })

@login_required(login_url='admin_login')
def admin_doctors(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    doctors = Doctor.objects.select_related('user').order_by('-id')

    return render(request, 'adminpanel/doctors/list.html', {
        'doctors': doctors
    })


@login_required(login_url='admin_login')
def admin_add_doctor(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    form = DoctorForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect('admin_doctors')

    return render(request, 'adminpanel/doctors/add.html', {
        'form': form
    })


@login_required(login_url='admin_login')
def admin_edit_doctor(request, doctor_id):
    if not request.user.is_superuser:
        return redirect('admin_login')

    doctor = get_object_or_404(Doctor, id=doctor_id)
    form = DoctorForm(request.POST or None, request.FILES or None, instance=doctor)

    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect('admin_doctors')

    return render(request, 'adminpanel/doctors/edit.html', {
        'form': form,
        'doctor': doctor
    })


@login_required(login_url='admin_login')
def admin_delete_doctor(request, doctor_id):
    if not request.user.is_superuser:
        return redirect('admin_login')

    doctor = get_object_or_404(Doctor, id=doctor_id)

    if request.method == "POST":
        doctor.delete()
        return redirect('admin_doctors')

    return render(request, 'adminpanel/doctors/delete.html', {
        'doctor': doctor
    })


def export_appointments_excel(request):
    """Export appointments as Excel file"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(
            "openpyxl is not installed. Please install it using: pip install openpyxl",
            status=400
        )
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Patient Name', 'Doctor', 'Appointment Date', 'Appointment Time', 'Status']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    appointments = Appointment.objects.all()
    for a in appointments:
        ws.append([
            a.id,
            a.patient_name,
            str(a.doctor),
            a.appointment_date.strftime('%d-%m-%Y') if a.appointment_date else '',
            a.appointment_time.strftime('%H:%M') if a.appointment_time else '',
            a.status
        ])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="appointments.xlsx"'
    wb.save(response)
    
    return response


@staff_member_required
def admin_delete_appointment(request, id):
    appointment = get_object_or_404(Appointment, id=id)
    appointment.delete()

    messages.success(request, "Appointment deleted successfully.")
    return redirect("admin_appointments")


@staff_member_required
def admin_feedback(request):
    feedbacks = ContactMessage.objects.all().order_by("-created_at")
    return render(request, "adminpanel/feedback.html", {
        "feedbacks": feedbacks
    })
